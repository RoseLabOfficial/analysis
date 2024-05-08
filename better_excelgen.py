import pandas as pd
import numpy as np
from dataclasses import dataclass
from openpyxl import load_workbook
import json

import os, re
from typing import List, Optional, Dict, Tuple, Set

def numeric(s: str) -> float | int:
    match = re.search(r'-?\d*\.?\d+', s)
    if match:
        return float(match.group()) if '.' in match.group() else int(match.group())
    return None

@dataclass(frozen=True)
class WavAvg:
    filename: str
    atten: float
    carrier_freq: float
    pulse_dur: float
    pulse_n: int
    pps: float
    iinj: float
    version: str
    is_stim: bool

    def new(filename: str) -> "WavAvg":
        fields: List[str] = os.path.splitext(os.path.basename(filename))[0].split('_')
        args = [numeric(i) for i in fields[:6]]
        return WavAvg(filename, *args, version=fields[6], is_stim=len(fields) >= 8)

class Experiment:
    def __init__(self, wavavg0: WavAvg, data0: pd.DataFrame) -> None:
        self.output_filename: float = f"{wavavg0.filename[:wavavg0.filename.find('pulses')+len('pulses')]}.xlsx"

        self.atten: float = wavavg0.atten
        self.carrier_freq: float = wavavg0.carrier_freq
        self.pulse_dur: float = wavavg0.pulse_dur
        self.pulse_n: float = wavavg0.pulse_n

        self.time: np.ndarray = data0['Time'].to_numpy()
        self.stimuli: Dict[Tuple[float, str], pd.DataFrame] = {}
        self.voltages: Dict[Tuple[float, str], pd.DataFrame] = {}

        self.add(wavavg0, data0)

    def compat(self, wavavg: "WavAvg") -> bool:
        if any((
            self.atten != wavavg.atten,
            self.carrier_freq != wavavg.carrier_freq,
            self.pulse_dur != wavavg.pulse_dur,
            self.pulse_n != wavavg.pulse_n
        )):
            return False
        return True

    def add(self, wavavg: WavAvg, data: pd.DataFrame) -> None:
        key: Tuple[float, str, float] = (wavavg.pps, wavavg.version, wavavg.iinj * 1e-9)
        if wavavg.is_stim:
            self.stimuli[key] = data['1 Stimulus'].to_numpy()
        else:
            self.voltages[key] = data['1 Spikes+hr'].to_numpy()

    def get_output_path(self, output_folder: str) -> str:
        return os.path.join(output_folder, self.output_filename)

    def get_all_pps(self) -> Set[float]:
        return set(key[0] for key in self.voltages.keys())

    def get_sheet_df(self, pps: float, version: Optional[str]=None) -> pd.DataFrame:
        def filter_by_pps_version(avgs: Dict[Tuple[float, str], pd.DataFrame]):
            return {key:label for key, label in avgs.items() if key[0] == pps and (key[1] == version if version else True)}
        return filter_by_pps_version(self.stimuli), filter_by_pps_version(self.voltages)

class ExcelGen:
    def __init__(self, params_template_path: str='./parameters.json') -> None:
        with open(params_template_path, 'r') as f:
            parms_dict: dict = json.load(f)
        self.params_template = pd.DataFrame(parms_dict)

    @staticmethod
    def _read_averages(txtfile_folder) -> List[Tuple[WavAvg, pd.DataFrame]]:
        return [
            (WavAvg.new(filename), pd.read_csv(os.path.join(txtfile_folder, filename), sep='\t', engine='python'))
            for filename in os.listdir(txtfile_folder) if filename.endswith('.txt')
        ]
    
    @staticmethod
    def _group_averages(averages: List[Tuple[WavAvg, pd.DataFrame]]) -> List[Experiment]:
        experiments: List[Experiment] = []
        for wavavg, datum in averages:
            new: bool = True
            for exp in experiments:
                if exp.compat(wavavg):
                    exp.add(wavavg, datum)
                    new = False
                    break
            if new:
                experiments.append(Experiment(wavavg, datum))
        return experiments

    @staticmethod
    def _ui_cond(v: str | List[str], versions: Set[str]) -> bool:
        if isinstance(v, list):
            return all([i in versions for i in v])
        return v in versions

    @staticmethod
    def _ui_inputfn(inputstr: str, qrytype: str) -> str | List[str]:
        if qrytype == 'stimulus':
            return inputstr
        return list(inputstr.split(','))

    @staticmethod
    def _user_select(dic: Dict[Tuple[float, str], np.ndarray], qrytype: str) -> Dict[Tuple[float, str], np.ndarray]:
        versions: Set[str] = set(key[1] for key in dic.keys())
        v: str = ExcelGen._ui_inputfn(
            input(f"Select {qrytype} version(s):\n"+[f"{key[1]}\n" for key in dic.keys()]).replace(' ', ''), qrytype
        )
        while True:
            if ExcelGen._ui_cond(v, versions):
                break
            v = input(f"invalid selection: {v}\n")
        return {key:label for key, label in dic if key[1] == v}

    def _write_averages(self, exp: Experiment, output_folder: str, version: Optional[str]) -> None:
        writer = pd.ExcelWriter(exp.get_output_path(output_folder), engine='openpyxl', mode='w')
        all_rates: Set[float] = exp.get_all_pps()
        for pps in sorted(all_rates):
            sheet_name: str = f"{pps}pps"
            print(f"\t\t{sheet_name}")

            stimulus, voltages = exp.get_sheet_df(pps, version)

            if not version:
                if len(stimulus) > 1:
                    stimulus = self._user_select(stimulus, 'stimulus')
                voltages = self._user_select(voltages, 'voltages')   

            print(voltages) 

            data: Dict[str, np.ndarray] = {'times': exp.time}
            data.update({key[2]:voltages[key] for key in sorted(voltages.keys(), key=lambda x: -x[2])})
            if len(stimulus) != 0:
                data.update({'stimulus': list(stimulus.values())[0]})

            df: pd.DataFrame = pd.DataFrame(data, )

            df.to_excel(writer, sheet_name=sheet_name, index=False)
            self.params_template.to_excel(writer, sheet_name=f"parameters_{sheet_name}", index=False)

        writer._save()

    @staticmethod
    def _write_parameters(exp: Experiment, output_folder: str) -> None:
        wb = load_workbook(exp.get_output_path(output_folder))

        def idx_to_str(i: int) -> str:
            return str(chr(i + 65))

        for sheet_name in sorted(wb.sheetnames, key=lambda x: float(''.join(i for i in x if i.isdigit()))):
            if 'parameters' in sheet_name:
                continue
            print(sheet_name)

            params_sheet = wb[f"parameters_{sheet_name}"]
            data_sheet = wb[sheet_name]

            j: int = 2
            for i, cell in enumerate(data_sheet[1]):
                if cell.value is None:
                    break
                if isinstance(cell.value, float) or isinstance(cell.value, int):
                    col: str = idx_to_str(i).upper()
                    params_sheet[f"A{j}"] = f"={sheet_name}!{col}1"
                    params_sheet[f"D{j}"] = f"=({sheet_name}!{col}2)*0.001"
                    params_sheet[f"H{j}"] = f"=I{j}+0.025"
                    params_sheet[f"I{j}"] = f"=(A{j}*C{j})+D{j}"
                    params_sheet[f"J{j}"] = 1
                    params_sheet[f"K{j}"] = 1
                    params_sheet[f"M{j}"] = f"=I{j}"
                    params_sheet[f"N{j}"] = float(sheet_name[:-3])
                    params_sheet[f"O{j}"] = exp.pulse_n
                    params_sheet[f"P{j}"] = f"=B{j}*C{j}"
                    
                    j += 1

        wb.save(exp.get_output_path(output_folder))

    def __call__(self, txtfile_folder: str, output_folder: str, version: Optional[str]) -> None:
        averages: Dict[WavAvg, pd.DataFrame] = self._read_averages(txtfile_folder)
        experiments: List[Experiment] = self._group_averages(averages)

        for exp in experiments:
            print(f"\n{exp.output_filename}\n\tWriting Averages...")
            self._write_averages(exp, output_folder, version)
            print("\tWriting Parameters...")
            self._write_parameters(exp, output_folder)

if __name__ == '__main__':
    import sys
    eg = ExcelGen(sys.argv[4]) if len(sys.argv) >= 5 else ExcelGen()
    eg(sys.argv[1], sys.argv[2], sys.argv[3] if len(sys.argv) >= 4 else None)